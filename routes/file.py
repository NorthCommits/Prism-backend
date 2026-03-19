import io
import csv
import openpyxl
from fastapi import APIRouter, UploadFile, File, HTTPException
from pydantic import BaseModel

router = APIRouter()


def parse_csv_content(content: bytes) -> str:
    try:
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return "Empty CSV file."

        result_lines = []
        headers = rows[0]
        result_lines.append(f"Columns: {', '.join(headers)}")
        result_lines.append(f"Total rows: {len(rows) - 1}")
        result_lines.append("")

        for i, row in enumerate(rows):
            result_lines.append(" | ".join(row))

        return "\n".join(result_lines)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"CSV parsing error: {str(e)}")


def parse_xlsx_content(content: bytes) -> str:
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
        result_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            result_parts.append(f"=== Sheet: {sheet_name} ===")

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                result_parts.append("(empty sheet)")
                continue

            result_parts.append(f"Total rows: {len(rows)}")
            result_parts.append("")

            for row in rows:
                cleaned = [str(cell) if cell is not None else "" for cell in row]
                result_parts.append(" | ".join(cleaned))

            result_parts.append("")

        return "\n".join(result_parts)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel parsing error: {str(e)}")


class ParsedFile(BaseModel):
    file_name: str
    file_type: str
    content: str


@router.post("/file/parse", response_model=ParsedFile)
async def parse_file(file: UploadFile = File(...)):
    file_name = file.filename or "unknown"
    extension = file_name.split(".")[-1].lower()

    allowed_extensions = ["txt", "md", "py", "js", "ts", "jsx", "tsx", "csv", "xlsx"]
    if extension not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '.{extension}'. Allowed: {allowed_extensions}"
        )

    content_bytes = await file.read()

    if extension == "csv":
        content = parse_csv_content(content_bytes)
        file_type = "csv"
    elif extension == "xlsx":
        content = parse_xlsx_content(content_bytes)
        file_type = "excel"
    else:
        # plain text / code files
        content = content_bytes.decode("utf-8", errors="replace")
        file_type = "code" if extension in ["py", "js", "ts", "jsx", "tsx"] else "text"

    # limit content to 50000 chars to avoid token overflow
    if len(content) > 50000:
        content = content[:50000] + "\n\n[Content truncated due to size limit]"

    return ParsedFile(file_name=file_name, file_type=file_type, content=content)